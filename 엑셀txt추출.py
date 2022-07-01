from openpyxl import Workbook,load_workbook

wb= load_workbook(r"C:\Users\Notebook\Downloads\--.xlsx")
ws=wb['Sheet2']
for x in range(2,33):
    if x==21 or x==30:
        continue
    ws['H{}'.format(x)]='안녕하세요! 만나서 반갑습니다. '+ws.cell(x,8).value

	
wb.save(r"C:\Users\Notebook\Downloads\--.xlsx") # 경로없이 이름만 쓰면 이 py파일이 있는 폴더에 저장됨
wb.close()
